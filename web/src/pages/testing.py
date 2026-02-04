import os
import time
from datetime import datetime
from playwright.sync_api import sync_playwright, TimeoutError
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# ────────────────────────────────────────────────
# CONFIG
# ────────────────────────────────────────────────
PROFILE_DIR = r"C:\\playwright\\bowens-profile"
BASE_URL = "https://www.bowens.com.au"
CATEGORY_URL = f"{BASE_URL}/c/timber/"
OUTPUT_DIR  = "."                              # where to save XLSX
HEADLESS    = True

# Tunable timings
APPEAR_TIMEOUT_MS = 4000   # wait for next dropdown to appear after a selection
PRICE_SETTLE_MS   = 4500   # wait for price to update
DROPDOWN_PAUSE_MS = 220    # short pause after selecting/resetting

# ────────────────────────────────────────────────
# HELPERS
# ────────────────────────────────────────────────
def safe_js_click(page, locator):
    page.evaluate("(el) => el.click()", locator)

def get_text_or_empty(locator):
    try:
        return (locator.inner_text() or "").strip()
    except Exception:
        return ""

def wait_for_price_settle(page, price_locator, timeout=PRICE_SETTLE_MS):
    try:
        page.wait_for_timeout(200)
        price_locator.wait_for(state="visible", timeout=timeout)
    except TimeoutError:
        pass

# ────────────────────────────────────────────────
# SCRAPING FUNCTIONS
# ────────────────────────────────────────────────
def collect_product_urls(page):
    urls = set()
    page_num = 1
    while True:
        if page_num == 1:
            page.goto(START_URL, wait_until="networkidle", timeout=30000)

        links = page.locator("a[href*='/p/']")
        try:
            links.first.wait_for(state="attached", timeout=5000)
        except TimeoutError:
            pass

        added = 0
        for i in range(links.count()):
            href = links.nth(i).get_attribute("href") or ""
            if "/p/" in href:
                full = BASE_URL + href if href.startswith("/") else href
                clean = full.split("?")[0].rstrip("/")
                if clean not in urls:
                    urls.add(clean)
                    added += 1
        print(f"  +{added} → {len(urls)} total (page {page_num})")

        next_btn = page.locator('[aria-label="Next"], .pagination__next, a:has-text("Next"), button:has-text("Next")')
        if next_btn.count() == 0 or not next_btn.first.is_enabled():
            break

        safe_js_click(page, next_btn.first)
        time.sleep(2.0)
        page_num += 1

    return list(urls)

def scrape_product(page, url, idx, total):
    print(f"[{idx}/{total}] {url}")
    page.goto(url, wait_until="networkidle", timeout=35000)
    page.wait_for_timeout(600)

    title = get_text_or_empty(page.locator("h1, .product-title, [data-product-title]"))
    price_locator = page.locator(".price, .product-price, [data-price], [itemprop='price']")

    records = []

    def capture(chosen_labels):
        sku_now = get_text_or_empty(page.locator(".sku, [itemprop='sku'], [data-sku]"))
        wait_for_price_settle(page, price_locator, timeout=PRICE_SETTLE_MS)
        price_text = get_text_or_empty(price_locator)
        records.append({
            "title": title,
            "variation": " | ".join(chosen_labels),
            "sku": sku_now,
            "price": price_text,
            "url": url
        })

    def dfs(depth, chosen_labels):
        dropdowns = page.locator("select:not([disabled])")
        drop_count = dropdowns.count()

        if drop_count <= depth or depth >= 4:  # cap at 4 dropdowns
            capture(chosen_labels)
            return

        dropdown = dropdowns.nth(depth)
        options = dropdown.locator("option:not([disabled])")
        opt_count = options.count()
        for oi in range(opt_count):
            val = options.nth(oi).get_attribute("value") or ""
            lbl = get_text_or_empty(options.nth(oi))
            if val.strip() == "" and lbl.strip() == "":
                continue

            selected = False
            for sel_kwargs in ({"value": val}, {"label": lbl}, {"index": oi}):
                try:
                    dropdown.select_option(**sel_kwargs)
                    selected = True
                    break
                except Exception:
                    continue
            if not selected:
                continue

            page.wait_for_timeout(DROPDOWN_PAUSE_MS)

            try:
                page.wait_for_function(
                    "(expected) => document.querySelectorAll('select:not([disabled])').length >= expected",
                    arg=drop_count + 1,
                    timeout=APPEAR_TIMEOUT_MS
                )
            except TimeoutError:
                pass

            dfs(depth + 1, chosen_labels + [lbl or val])

            # Reset only this dropdown; parents remain selected
            try:
                dropdown.select_option(index=0)
            except Exception:
                pass
            page.wait_for_timeout(DROPDOWN_PAUSE_MS)

    first_dropdown = page.locator("select:not([disabled])")
    if first_dropdown.count() == 0:
        capture([])
    else:
        dfs(0, [])

    return records

# ────────────────────────────────────────────────
# EXCEL EXPORT
# ────────────────────────────────────────────────
def save_to_excel(data, filename):
    wb = Workbook()
    ws = wb.active
    ws.title = "Bowens Timber"

    headers = ["Product", "Variation", "SKU", "Price", "URL"]
    ws.append(headers)

    for item in data:
        ws.append([item['title'], item['variation'], item['sku'], item['price'], item['url']])

    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for col in ws.columns:
        max_length = 0
        column = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[column].width = min(max_length + 3, 70)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    wb.save(filename)
    print(f"✓ Saved {len(data)} rows → {filename}")

# ────────────────────────────────────────────────
# MAIN
# ────────────────────────────────────────────────
def main():
    print("Bowens Timber Scraper v3 – Cascading Variants\n" + "="*50)

    with sync_playwright() as p:
        context = p.chromium.launch_persistent_context(
            PROFILE_DIR,
            headless=HEADLESS,
            viewport={"width": 1400, "height": 900},
            ignore_https_errors=True,
        )
        page = context.new_page()

        print("\nCollecting product URLs...")
        product_urls = collect_product_urls(page)
        print(f"→ {len(product_urls)} products found")

        all_data = []
        for idx, url in enumerate(product_urls, 1):
            if idx % 80 == 0 and idx < len(product_urls):
                print("  🔄 Restarting browser context...")
                context.close()
                context = p.chromium.launch_persistent_context(
                    PROFILE_DIR,
                    headless=HEADLESS,
                    viewport={"width": 1400, "height": 900},
                    ignore_https_errors=True,
                )
                page = context.new_page()

            products = scrape_product(page, url, idx, len(product_urls))
            all_data.extend(products)

            if idx % 15 == 0:
                print(f"  Progress: {idx}/{len(product_urls)} | {len(all_data)} variants total")
                ts = datetime.now().strftime("%Y%m%d_%H%M%S")
                os.makedirs(OUTPUT_DIR, exist_ok=True)
                filename = os.path.join(OUTPUT_DIR, f"bowens_timber_{ts}.xlsx")
                save_to_excel(all_data, filename)

        print("\nComplete.")
        input("Press Enter to close browser...")
        context.close()

    if name == "main":
        main()  
